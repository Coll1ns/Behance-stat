import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from datetime import date
import openpyxl
data = {}


# function that is getting values from the behance site
def get_stat_today():
    html = requests.get('https://www.behance.net/tanya_durs').text
    soup = BeautifulSoup(html, 'html.parser')
    value = soup.find_all('tr', class_="UserInfo-statRow-Erw")
    key = 0

    for element in value:
        # key = element.find('td', class_='UserInfo-statColumn-1vg').text
        key_val = element.find(
            'td', class_='UserInfo-statColumn-1vg UserInfo-statValue-1_-').text
        key_val = key_val.replace(',', '')
        data[key] = int(key_val)
        key += 1


def get_diff():
    # df = pd.read_excel('Stats.xlsx')
    # reading and flipping the table to extract previouse day data
    df = pd.read_excel("test.xlsx")
    df = df.T

    # getting the last day data and making the diff arg that will be used to form the
    # new row of data
    last_day_stat = df.loc[df.index[-1], 0:3]
    last_day_stat = pd.to_numeric(last_day_stat)
    current_day_stat = pd.Series(data)
    diff = current_day_stat - last_day_stat

    # here we are creating a column that we are going to add to the table
    infront = pd.Series("-", index=[4])
    current_day_stat = current_day_stat.append([infront, diff], ignore_index=True)
    current_day_stat = current_day_stat.values.tolist()
    current_day_stat.insert(0, date.today())
    return current_day_stat


def add_column(sheet_name, column):
    ws = wb[sheet_name]
    new_column = ws.max_column + 1

    for rowy, value in enumerate(column, start = 1):
        ws.cell(row = rowy, column = new_column, value = value)


get_stat_today()

wb = openpyxl.load_workbook('test.xlsx')
add_column('days', get_diff())
wb.save('test.xlsx')